from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import *
from PyQt5.QtWidgets import QMessageBox
import os
import time
#import xlsxwriter
import openpyxl
class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.setFixedSize(160, 190)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.lineEdit = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit.setGeometry(QtCore.QRect(10, 10, 141, 31))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("İcon Dosya yolu"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        MainWindow.setWindowIcon(icon)
        self.lineEdit.setFont(font)
        self.lineEdit.setObjectName("lineEdit")
        self.lineEdit_2 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_2.setGeometry(QtCore.QRect(10, 50, 141, 31))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_2.setFont(font)
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.lineEdit_3 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_3.setGeometry(QtCore.QRect(10, 90, 141, 31))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_3.setFont(font)
        self.lineEdit_3.setObjectName("lineEdit_3")
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(30, 130, 101, 51))
        font = QtGui.QFont()
        font.setFamily("Bookman Old Style")
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton.setFont(font)
        self.pushButton.setObjectName("pushButton")
        MainWindow.setCentralWidget(self.centralwidget)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.lineEdit.setPlaceholderText(_translate("MainWindow", "MÜŞTERİ ADI"))
        self.lineEdit_2.setPlaceholderText(_translate("MainWindow", "SİPARİŞ NO"))
        self.lineEdit_3.setPlaceholderText(_translate("MainWindow", "İŞ EMRİ NO"))
        self.pushButton.setText(_translate("MainWindow", "ETİKET \n" "OLUŞTUR"))
        MainWindow.setStyleSheet("background-color:GhostWhite; border-radius:8px")
        self.lineEdit.setStyleSheet(" border: 1px solid black ; border-radius:8px")
        self.lineEdit_2.setStyleSheet(" border: 1px solid black ; border-radius:8px")
        self.lineEdit_3.setStyleSheet(" border: 1px solid black ; border-radius:8px")
        self.pushButton.setStyleSheet("background-color:LightGreen; border: 1px solid black ; border-radius:6px")
        self.pushButton.clicked.connect(self.etiket)
    def etiket(self):
        import pandas as pd
        musteri = (self.lineEdit.text())
        siparis = (self.lineEdit_2.text())
        isemri = (self.lineEdit_3.text())
        print(musteri, " ", siparis, " ", isemri)
        hata = False
        excel_file = r".../FormS.xlsx"
        
        df = pd.read_excel(excel_file)
       

        
        if musteri == "":
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            aa = "'Müşteri adı' kısmına, Projenin ait olduğu müşteri ismini girmelisin."
            msg.setText(aa)
            msg.setWindowTitle("Boş bıraktın")
            retval = msg.exec_()
            hata = True
        elif siparis == "":
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            aa = "'Sipariş No' kısmına, Projenin/ Siparişin ismini veya numarasını  girmelisin."
            msg.setText(aa)
            msg.setWindowTitle("Boş bıraktın")
            retval = msg.exec_()
            hata = True

        elif isemri == "":
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            aa = "'İş Emri No' kısmına, Punch Takip Listesinden alınan sıra numarasını girmelisin"
            msg.setText(aa)
            msg.setWindowTitle("Boş bıraktın")
            retval = msg.exec_()
            hata = True
        elif hata != True:    
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            aa = "Masaüstünüzde bulunan 'Etiket' adlı klasörün içindeki " + siparis + " projelerine ait setuplardan etiket oluşturulacaktır. \n \nHazırsa devam edin"
            msg.setText(aa)
            msg.setWindowTitle("UYARI")
            hata = True
            retval = msg.exec_()
        excelyol = r"...\ETIKET\ETIKET"
        exceller = os.listdir(excelyol)
        etiketsay = 0
        MainWindow.hide()
        
        for setup in exceller:
            if setup.endswith(".xlsx"):
                exsetup_1 = load_workbook(excelyol + "\\" + setup)
                
                etiketexcel_1 = load_workbook(r"...\ETSBLN.xlsx")
                bukumexcel_1 = load_workbook(excel_file)
                
                #sheet = etiketexcel['Sayfa1']
                setupadi = str(setup)
                ien = setupadi[-8:-5]
                print("SETUP ADI: " , setup)
                print("\n--------\n")
                SAY = 9
                bukumexcel =bukumexcel_1.active
                exsetup=exsetup_1.active
                
                a = True
                while a == True:
                    SAYSTR = str(SAY)
                    kod= str(exsetup["B" + SAYSTR].value)
                    adet= str(exsetup["H" + SAYSTR].value)
                    plaka= str(exsetup["A" + "1"].value)
                    
                    print("Tam hali: " , kod)
                    kod = kod.replace(".dft", ".DFT")
                    kodindex = kod.find(".DFT")
                    kodindex = kodindex- 1
                    #print(kodindex)
                    if kodindex != -2:
                        
                        kod92mi = kod[0:2]
                        if kod92mi == "92":
                            SL = r".../92.xlsx"
                            siparis = "92"
                            '''
                            aranacakkodstr = kod[0:10]
                            '''
                            kod = kod[3:10]
                            
                        else:
                            siparis = "91"
                            SL = r".../91.xlsx"
                            kod = kod[0:7]
                        siparislistesi_1 = load_workbook(SL)
                        siparislistesi = siparislistesi_1.active
                        aranacakkodstr = int(kod)
                        sloku = pd.read_excel(SL)    
                        print("Sorgu sonrası:",kod)
                        print("ARANACAK KOD",aranacakkodstr)
                        time.sleep(1)
                        
                        kodi = int(kod)
                        #print(kodi)
                        SAY = SAY + 1
                        tpkod = (df["BS"].where(df["PK"]== kodi))
                        tpbuk = tpkod.dropna()
                        tpbuk = str(tpbuk)
                        bosluk = tpbuk.find(" ")
                        kodistr = str(kodi)
                        print(kodi)
                        print(tpbuk)
                        if tpbuk[4] != " ":
                            msg = QMessageBox()
                            msg.setIcon(QMessageBox.Information)
                            aa = kodistr + " nolu parçanın bükümü varmı ?\nExcele girin."
                            msg.setText(aa)
                            msg.setWindowTitle("UYARI")
                            retval = msg.exec_()
                            break
                        toplambuk = int(tpbuk[0:bosluk])
                        toplambuk = str(toplambuk + 2)
                        print("BUKUM " + toplambuk)
                        time.sleep(1)
                        tpslL = (sloku["ADET"].where(sloku["KOD"]== aranacakkodstr))
                        #print("tpslL", tpslL)
                        tpsl = tpslL.dropna()
                        tpsl = str(tpsl)
                        bosluk = tpsl.find(" ")
                        print("tpsl", tpsl)
                        toplamsiparis = int(tpsl[0:bosluk])
                        toplamsiparis = str(toplamsiparis + 2)
                        
                        siparisadedi = str(siparislistesi["B" + toplamsiparis].value)
                        
                        abkant =  (bukumexcel["B" + toplambuk].value)
                        time.sleep(1)
                        boya = (sloku["BOYA"].where(sloku["KOD"]== aranacakkodstr))
                        
                        boya = boya.dropna()
                        boya = str(boya)
                        bosluk = tpsl.find(" ")
                        boyaindex = int(boya[0:bosluk])
                        boyaindex = str(boyaindex + 2)
                        ral =  (siparislistesi["C" + boyaindex].value)
                        boya = ral.replace("RAL", "")
                        print("BOYA KODU",boya)
                        print(toplamsiparis)
                        etiketsay = etiketsay + 1
                        
                        etiketsaystr = str(etiketsay)
                        #print(abkant)
                        etiket=etiketexcel_1.active
                        if etiketsay < 10:
                            etiketsaystr = "00" + etiketsaystr
                        if 9 < etiketsay and adetsay < 100 :
                            etiketsaystr = "0" + etiketsaystr
                        if 99 < etiketsay:
                            etiketsaystr = str(etiketsay)
                        
                        ien = str(ien)
                        isem = isemri + " - " + ien 
                        yaziyeni = ""
                        
                        for harf in kodistr:
                            
                            yaziyeni = yaziyeni + "  " + harf
                        
                        
                        adetsay=1
                        plakaint = int(plaka)
                        adetint = int(adet)
                        adetint = adetint * plakaint
                        
                        
                        # netsis = r"...\NETSIS.xlsx"
                        # netsisoku = pd.read_excel(netsis)   
                        # netsislistesi_1 = openpyxl.load_workbook(netsis)
                        # netsislistesi = netsislistesi_1.active
                        # netsiskodu = (netsisoku["Stok Kodu "].where(netsisoku["PARCAADI"]== aranacakkodstr))
                        # print(netsiskodu)
                        

                        # nkodu = netsiskodu.dropna()
                        # nkodu = str(nkodu)
                        # bosluk = nkodu.find(" ")
                        # nkoduindex = int(nkodu[0:bosluk])
                        # nkoduindex = str(nkoduindex + 2)
                        
                        # time.sleep(1)
                        # nk = (netsislistesi["A" + nkoduindex].value)
                        # print(nk)
                        
                        while adetsay <= adetint :
                            if abkant != 0:
                                etiket['B8'] = "X"
                            #if ral != "NO":
                            time.sleep(1)
                            etiket['J7'] = boya
                            etiket['G2'] = yaziyeni
                            #etiket['G3'] = nk
                            etiket['G3'] = musteri[0] + kodistr
                            etiket['D2'] = musteri
                            etiket['D3'] = siparis
                            etiket['B4'] = isemri + " - "
                            etiket['D4'] = ien
                            adetsaystr = str(adetsay)
                            if adetsay < 10:
                                adetsaystr = "00"+adetsaystr
                            if 9 < adetsay and adetsay < 100 :
                                adetsaystr = "0"+adetsaystr
                            if 99 < adetsay:
                                adetsaystr = str(adetsay)
                                
                            etiket['J4'] = " / " + siparisadedi
                            
                            #print(kod + adet)
                            
                            # etiketyazi = "ETIKET" + etiketsaystr + "-"
                            etiketyazi = kodistr  + "-" + etiketsaystr
                            # etiketexcel.save(excelyol + "\\" + etiketyazi + setup )
                            #try:
                            etiket['C12'] = etiketyazi + "-" + adetsaystr +".xlsx"
                            #except:
                            #continue
                            print(etiketyazi + "-" + adetsaystr +".xlsx")
                            print("\n--------\n")
                            exsetup_1.close()
                           
                            bukumexcel_1.close()
                           
                            siparislistesi_1.close()
                            #etiketexcel_1.save(excelyol + "\\ETIKETLER\\" + etiketyazi + ".xlsx")
                            ws = etiketexcel_1.worksheets[0]
                            etiketexcel_1.close()
                            etiket=etiketexcel_1.active
                            time.sleep(0.2)
                            img = openpyxl.drawing.image.Image("LOTUS11.jpg")
                            img.width = 38
                            img.height = 38
                            img.anchor = 'A1'
                            ws.add_image(img)
                            try:
                                
                                os.mkdir(excelyol + "\\ETIKETLER\\"+ siparis + "-" + kodistr)
                                etiketexcel_1.save(excelyol + "\\ETIKETLER\\"+ siparis + "-" + kodistr + "\\" + etiketyazi + "-" + adetsaystr +".xlsx")
                            except:
                                etiketexcel_1.save(excelyol + "\\ETIKETLER\\" + siparis + "-" + kodistr + "\\" + etiketyazi + "-" + adetsaystr +".xlsx")
                                
                            
                            etiketexcel_1.close()
                            adetsay += 1
                            '''
                            workbook = xlsxwriter.Workbook(excelyol + "\\ETIKETLER\\" + etiketyazi + ".xlsx")
                            worksheet = workbook.add_worksheet()
                            worksheet.insert_image('A1', 'LOTUS.png', {'x_offset': 15, 'y_offset': 10})
                            workbook.close()
                            '''
                        
                    if kod == "SON":
                        kod = ""
                        a =False
        dosyayolu = os.listdir(r"...\ETIKET\ETIKET\ETIKETLER")
        kacoldu = 0
        for kods in dosyayolu:
            
            kodyol = os.listdir("...\ETIKET\ETIKET\ETIKETLER" + "\\" + kods)
            say = 1
            for etikety in kodyol:
                time.sleep(0.2)
                print(etikety)
                exetiket_1 = load_workbook("...\ETIKET\ETIKET\ETIKETLER" + "\\" + kods + "\\" + etikety)
                exetiket=exetiket_1.active
                saystr = str(say)
                SAYAC = str(exetiket["J" + "4"].value)
                
                
                SAYACADET = SAYAC.replace(" / ", "")
                
                print(saystr + SAYAC)
                siparisadediint = int(SAYACADET)
                print(SAYACADET)
                if say > siparisadediint:
                    exetiket['J4'] = "STOK"
                else:
                    exetiket['J4'] = saystr + SAYAC
                say += 1
                time.sleep(0.2)
                kacoldu += 1
                kacoldustr = str(kacoldu)
                exetiket_1.save("...\ETIKET\ETIKET\ETIKETLER" + "\\" + kods + "\\" + etikety)
                exetiket_1.close()
                
        for kods in dosyayolu:
            kodyol = os.listdir("...\ETIKET\ETIKET\ETIKETLER" + "\\" + kods)
            time.sleep(0.2)
            for etikety in kodyol:
                
                    
                time.sleep(0.2)
                print(etikety)
                exetiket_1 = load_workbook("...\ETIKET\ETIKET\ETIKETLER" + "\\" + kods + "\\" + etikety)
                exetiket=exetiket_1.active
                SAYAC = str(exetiket["D" + "4"].value)
                print(SAYAC)
                try:
                    os.mkdir(r".../ETIKET/ETIKETYERLESIM/" + SAYAC)
                    
                    os.rename(".../ETIKET/ETIKET/ETIKETLER" + "\\" + kods + "\\" + etikety
                        ,r".../ETIKET" + "/ETIKETYERLESIM/"+ SAYAC + "/"  + etikety)
                    
                except:
                    
                    os.rename(".../ETIKET/ETIKET/ETIKETLER" + "\\" + kods + "\\" + etikety
                        ,r".../ETIKET" + "/ETIKETYERLESIM/"+ SAYAC + "/" + etikety)
                
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        
        aa = kacoldustr + " adet etiket hazır"
        msg.setText(aa)
        msg.setWindowTitle("Bitti")
        retval = msg.exec_()
        dosyayoluu = os.listdir("...\ETIKET\ETIKET\ETIKETLER")
        for kodu in dosyayoluu:
           
            os.rmdir(".../ETIKET/ETIKET/ETIKETLER/" + kodu )               
                    
                
        # setup=load_workbook(excelyol)
        # setup1=setup.active
        # print(excelyol)
    
if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())